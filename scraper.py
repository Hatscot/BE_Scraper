"""
BrickEconomy Scraper
====================
Scrapet alle LEGO Sets von brickeconomy.com und exportiert sie als Excel-Datei.

Spalten:
  - Set Gruppe   (Theme, z.B. "City")
  - Set Nummer   (z.B. "60508")
  - Set Name     (als klickbarer Hyperlink zur BrickEconomy-Seite)
  - Jahr         (z.B. 2026)
  - Value Preis  (Marktwert laut BrickEconomy "Set Pricing"-Karte, z.B. "€199.99")
  - Growth       (Prozentuale Wertentwicklung, z.B. "+8.13%"), farblich markiert

Excel-Aufbau:
  - Ein einziges Sheet "LEGO Sets"
  - Jede Theme-Gruppe hat eine eigene farbige Überschriftszeile (wie eine "Mappe")
  - Datenzeilen mit alternierendem Hintergrund innerhalb jeder Gruppe
  - Growth-Spalte: Grün = positiv, Rot = negativ

Voraussetzungen:
  pip install requests beautifulsoup4 openpyxl lxml cloudscraper
"""

import re
import math
import time
import random
import logging
from datetime import datetime
from dataclasses import dataclass, field

import cloudscraper
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Logging ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("be_scraper")

# ─── Konfiguration ────────────────────────────────────────────────────────────
BASE_URL    = "https://www.brickeconomy.com"
SETS_URL    = f"{BASE_URL}/sets"
DELAY_PAGE  = (2.0, 5.0)    # Zufällige Pause zwischen Seiten eines Themes
DELAY_THEME = (5.0, 12.0)   # Zufällige Pause zwischen verschiedenen Themes
OUTPUT_DIR  = "D:\Entwicklung\GitHub\BE_Scraper_list"


# ─── Datenmodell ──────────────────────────────────────────────────────────────
@dataclass
class LegoSet:
    gruppe:  str
    nummer:  str
    name:    str
    url:     str   # Vollständige URL zur Set-Detailseite
    jahr:    str
    preis:   str   # Value-Preis
    growth:  str   # Growth in Prozent, z.B. "+8.13%" oder "-3.20%"


# ─── Hilfsfunktionen ─────────────────────────────────────────────────────────
def human_pause(min_s: float, max_s: float) -> None:
    delay = round(random.uniform(min_s, max_s), 2)
    log.debug("Warte %.2f Sekunden …", delay)
    time.sleep(delay)


def make_scraper():
    scraper = cloudscraper.create_scraper(
        browser={"browser": "chrome", "platform": "windows", "mobile": False}
    )
    scraper.headers.update({
        "Accept-Language": "de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept":
            "text/html,application/xhtml+xml,application/xml;q=0.9,"
            "image/avif,image/webp,*/*;q=0.8",
        "Referer": "https://www.brickeconomy.com/",
        "DNT": "1",
    })
    return scraper


def get_soup(url: str, session) -> BeautifulSoup:
    log.debug("GET %s", url)
    resp = session.get(url, timeout=30)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "lxml")


# ─── Theme-Liste ──────────────────────────────────────────────────────────────
def scrape_themes(session) -> list[dict]:
    log.info("Lade Themenliste von %s …", SETS_URL)
    soup = get_soup(SETS_URL, session)
    themes = []
    seen_urls = set()

    for a in soup.select("a[href*='/sets/theme/']"):
        href = a.get("href", "")
        if re.match(r"^/sets/theme/[^/\?]+$", href):
            name = a.get_text(strip=True)
            full_url = BASE_URL + href
            if name and full_url not in seen_urls:
                seen_urls.add(full_url)
                themes.append({"name": name, "url": full_url})

    log.info("  → %d Themen gefunden", len(themes))
    return themes


# ─── Value + Growth aus td.ctlsets-right ─────────────────────────────────────
def parse_value_and_growth(right_td) -> tuple[str, str]:
    """
    Liest Value-Preis und Growth-% aus der rechten Tabellenzelle der Listenansicht.

    Bestätigte HTML-Struktur in td.ctlsets-right (BrickEconomy):
      - Retail-Preis steht in <b>€XX.XX</b>
      - Value-Preis steht in <span class="cursor-help">€XX.XX</span>
        direkt nach dem Text-Label "Value"
      - Growth steht als Text-Node nach dem Label "Growth",
        davor oft ein <i class="icon-up-green-10"> oder "icon-down-red-10"
        Format: "+XX.X%" oder "-XX.X%"

    Fallback: Wenn kein Value-Preis existiert → Retail-Preis verwenden.
    """
    if not right_td:
        return "", ""

    value_price = ""
    growth_pct  = ""

    # ─── 0. PRIMÄR: div.ctlsets-value enthält direkt den Value-Preis ──────────
    # Bestätigte Struktur: <div class="ctlsets-value">Value <span class="cursor-help">€XX.XX</span></div>
    value_div = right_td.find("div", class_="ctlsets-value")
    if value_div:
        cursor_span = value_div.find("span", class_="cursor-help")
        if cursor_span:
            price_text = cursor_span.get_text(strip=True)
            if re.search(r"[€$£¥]\s*[\d,]+\.?\d*", price_text):
                value_price = price_text
        if not value_price:
            # Fallback: erster Preis im div
            pm = re.search(r"[€$£¥]\s*[\d,]+\.?\d*", value_div.get_text())
            if pm:
                value_price = pm.group(0).strip()

    # ─── 1. Value-Preis: span.cursor-help nach "Value"-Label ─────────────────
    # Iteriere über alle Kinder-Nodes um "Value"-Label → span.cursor-help zu finden
    children = list(right_td.descendants)
    for i, node in enumerate(children):
        # Suche Text-Node oder Element, das "Value" enthält (aber nicht "Retail")
        node_text = node.get_text(strip=True) if hasattr(node, "get_text") else str(node).strip()
        if node_text.lower() == "value":
            # Suche ab hier den nächsten span.cursor-help
            for j in range(i + 1, min(i + 15, len(children))):
                candidate = children[j]
                if hasattr(candidate, "get") and "cursor-help" in (candidate.get("class") or []):
                    price_text = candidate.get_text(strip=True)
                    if re.search(r"[€$£¥]\s*[\d,]+\.?\d*", price_text):
                        value_price = price_text
                        break
            if value_price:
                break

    # ─── 2. Fallback Value: Regex "Value ... €XX" im Plaintext ───────────────
    if not value_price:
        plain = right_td.get_text(separator="|", strip=True)
        m = re.search(r"Value[^€$£¥\d]{0,30}([€$£¥]\s*[\d,]+\.?\d*)", plain, re.I)
        if m:
            value_price = m.group(1).strip()

    # ─── 3. Fallback Value: zweiter Preis in der Zelle = Value ───────────────
    if not value_price:
        plain = right_td.get_text(separator="|", strip=True)
        all_prices = re.findall(r"[€$£¥]\s*[\d,]+\.?\d*", plain)
        if len(all_prices) >= 2:
            value_price = all_prices[1]   # Erster = Retail, zweiter = Value

    # ─── 4. Letzter Fallback: Retail-Preis aus <b> ───────────────────────────
    # (tritt auf wenn Set noch keinen Marktwert hat → aktive Sets)
    if not value_price:
        bold = right_td.find("b")
        if bold:
            value_price = bold.get_text(strip=True)
        if not value_price:
            pm = re.search(r"[€$£¥]\s*[\d,]+\.?\d*", right_td.get_text())
            if pm:
                value_price = pm.group(0).strip()

    # ─── 5a. Growth PRIMÄR: <i class="icon-up/down-..."> gefolgt von "%"-Text ─
    # Bestätigte Struktur: Growth <i class="icon-up-green-10"></i> +819.6%
    growth_icon = right_td.find("i", class_=re.compile(r"icon-(up|down)"))
    if growth_icon:
        # Text-Node nach dem Icon suchen
        for sibling in growth_icon.next_siblings:
            sib_text = sibling.get_text(strip=True) if hasattr(sibling, "get_text") else str(sibling).strip()
            m = re.search(r"([+\-]\s*[\d,.]+\s*%)", sib_text)
            if m:
                growth_pct = m.group(1).strip().replace(" ", "")
                break
        # Auch den Parent-Text prüfen, falls der Prozentwert ein Geschwister-Text ist
        if not growth_pct:
            parent = growth_icon.parent
            if parent:
                parent_text = parent.get_text(separator=" ", strip=True)
                m = re.search(r"([+\-]\s*[\d,.]+\s*%)", parent_text)
                if m:
                    growth_pct = m.group(1).strip().replace(" ", "")

    # ─── 5b. Growth: Text nach "Growth"-Label (Descendants scannen) ───────────
    if not growth_pct:
        children = list(right_td.descendants)
        for i, node in enumerate(children):
            node_text = node.get_text(strip=True) if hasattr(node, "get_text") else str(node).strip()
            if node_text.lower() == "growth":
                for j in range(i + 1, min(i + 20, len(children))):
                    candidate = children[j]
                    cand_text = candidate.get_text(strip=True) if hasattr(candidate, "get_text") else str(candidate).strip()
                    m = re.search(r"([+\-]\s*[\d,.]+\s*%)", cand_text)
                    if m:
                        growth_pct = m.group(1).strip().replace(" ", "")
                        break
                if growth_pct:
                    break

    # ─── 6. Fallback Growth: Regex "Growth ... +/-X%" im Plaintext ───────────
    if not growth_pct:
        plain = right_td.get_text(separator="|", strip=True)
        m = re.search(r"Growth[^+\-\d]{0,30}([+\-]\s*[\d,.]+\s*%)", plain, re.I)
        if m:
            growth_pct = m.group(1).strip().replace(" ", "")

    # ─── 7. Letzter Fallback: beliebiges +/-X%-Vorkommen ─────────────────────
    if not growth_pct:
        plain = right_td.get_text(separator=" ", strip=True)
        m = re.search(r"([+\-]\s*[\d,.]+\s*%)", plain)
        if m:
            growth_pct = m.group(1).strip().replace(" ", "")

    return value_price, growth_pct


# ─── Set-Zeilen parsen ────────────────────────────────────────────────────────
def parse_set_rows(soup: BeautifulSoup, theme_name: str) -> list[LegoSet]:
    sets = []

    table = soup.select_one("table.ctlsets-table")
    if not table:
        log.warning("Keine ctlsets-table gefunden – Seitenstruktur geändert?")
        return sets

    for row in table.find_all("tr"):
        left_td = row.find("td", class_="ctlsets-left")
        if not left_td:
            continue
        try:
            # ── Set-Link, -Nummer und -Name ───────────────────────────────────
            title_tag = left_td.select_one("h4 a")
            if not title_tag:
                continue
            title_text = title_tag.get_text(strip=True)

            # Set-URL
            href = title_tag.get("href", "")
            set_url = (BASE_URL + href) if href.startswith("/") else href

            # Format: "60508 Police Train Heist"
            match = re.match(r"^(\d[\d\-]*)[\s\u00a0]+(.+)$", title_text)
            if match:
                nummer = match.group(1).strip()
                name   = match.group(2).strip()
            else:
                nummer = ""
                name   = title_text

            # ── Jahr ──────────────────────────────────────────────────────────
            jahr = ""
            year_a = left_td.find("a", href=re.compile(r"/sets/year/(\d{4})"))
            if year_a:
                m = re.search(r"/sets/year/(\d{4})", year_a["href"])
                if m:
                    jahr = m.group(1)
            if not jahr:
                yt = re.search(r"Year\D{0,5}(\d{4})", left_td.get_text())
                if yt:
                    jahr = yt.group(1)

            # ── Value-Preis & Growth ──────────────────────────────────────────
            right_td = row.find("td", class_="ctlsets-right")
            value_price, growth_pct = parse_value_and_growth(right_td)

            # Wenn kein Value gefunden → Retail als Fallback (besser als leer)
            if not value_price and right_td:
                bold = right_td.find("b")
                if bold:
                    value_price = bold.get_text(strip=True)
                if not value_price:
                    pm = re.search(r"[€$£¥]\s*[\d,]+\.?\d*", right_td.get_text())
                    if pm:
                        value_price = pm.group(0).strip()

            sets.append(LegoSet(
                gruppe=theme_name,
                nummer=nummer,
                name=name,
                url=set_url,
                jahr=jahr,
                preis=value_price,
                growth=growth_pct,
            ))

        except Exception as exc:
            log.warning("Fehler beim Parsen einer Set-Zeile: %s", exc)

    return sets


# ─── Seitenanzahl ─────────────────────────────────────────────────────────────
def get_total_pages(soup: BeautifulSoup, page_size: int = 50) -> int:
    pager_text = soup.get_text()
    m = re.search(r"\d+\s+to\s+\d+\s+of\s+([\d,]+)\s+set", pager_text, re.I)
    if m:
        total_sets = int(m.group(1).replace(",", ""))
        table = soup.select_one("table.ctlsets-table")
        actual_size = 0
        if table:
            actual_size = len([r for r in table.find_all("tr")
                               if r.find("td", class_="ctlsets-left")])
        actual_size = actual_size or page_size
        return math.ceil(total_sets / actual_size)
    return 1


# ─── Theme scrapen ────────────────────────────────────────────────────────────
def scrape_theme(theme: dict, session) -> list[LegoSet]:
    all_sets: list[LegoSet] = []
    page = 1

    while True:
        url = f"{theme['url']}?CTY=1&O=0&page={page}"
        log.info("  Seite %d: %s", page, url)
        try:
            soup = get_soup(url, session)
        except Exception as e:
            log.error("Fehler bei %s: %s", url, e)
            break

        sets_on_page = parse_set_rows(soup, theme["name"])
        if not sets_on_page:
            log.info("  Keine Sets auf Seite %d → beende Theme", page)
            break

        all_sets.extend(sets_on_page)
        log.info("  → %d Sets geladen (gesamt: %d)", len(sets_on_page), len(all_sets))

        total_pages = get_total_pages(soup)
        if page >= total_pages:
            break

        page += 1
        human_pause(*DELAY_PAGE)

    return all_sets


# ─── Excel schreiben ──────────────────────────────────────────────────────────
def write_excel(all_sets: list[LegoSet], filepath: str) -> None:
    log.info("Schreibe Excel-Datei: %s", filepath)

    # ── Farb-Konstanten ───────────────────────────────────────────────────────
    HEADER_BG       = "1B4F8C"   # Dunkles Blau für Haupt-Header
    HEADER_FONT     = "FFFFFF"
    GROUP_BG        = "2E75B6"   # Mittleres Blau für Theme-Gruppen-Kopfzeilen
    GROUP_FONT      = "FFFFFF"
    ROW_ALT         = "EAF1FB"   # Helles Blau für alternierende Zeilen
    BORDER_COLOR    = "CCCCCC"
    GREEN_BG        = "C6EFCE"   # Excel-Standard Grün für positive Werte
    GREEN_FONT      = "276221"
    RED_BG          = "FFC7CE"   # Excel-Standard Rot für negative Werte
    RED_FONT        = "9C0006"
    NEUTRAL_BG      = "FFEB9C"   # Gelb für 0%
    NEUTRAL_FONT    = "9C6500"

    thin   = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LEGO Sets"

    # ── Spaltenbreiten ────────────────────────────────────────────────────────
    # Spalten: Set Gruppe | Set Nummer | Set Name | Jahr | Value | Growth
    col_widths = {1: 22, 2: 14, 3: 48, 4: 10, 5: 14, 6: 12}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Haupt-Header (Zeile 1) ────────────────────────────────────────────────
    headers = ["Set Gruppe", "Set Nummer", "Set Name", "Jahr", "Value (€)", "Growth"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = Font(bold=True, color=HEADER_FONT, size=11)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[1].height = 24

    # ── Sets nach Theme gruppieren ────────────────────────────────────────────
    from itertools import groupby
    groups = []
    for theme_name, group_sets in groupby(all_sets, key=lambda s: s.gruppe):
        groups.append((theme_name, list(group_sets)))

    current_row = 2

    for theme_name, theme_sets in groups:
        # ── Theme-Gruppen-Kopfzeile (Mappe) ──────────────────────────────────
        ws.row_dimensions[current_row].height = 20
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.fill   = PatternFill("solid", fgColor=GROUP_BG)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        # Theme-Name + Anzahl Sets
        header_cell = ws.cell(
            row=current_row, column=1,
            value=f"  📦 {theme_name}  ({len(theme_sets)} Sets)"
        )
        header_cell.font = Font(bold=True, color=GROUP_FONT, size=11)
        # Zellen zusammenführen (Merge A bis F für die Gruppenzeile)
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=6
        )
        current_row += 1

        # ── Datenzeilen für dieses Theme ─────────────────────────────────────
        for alt_idx, s in enumerate(theme_sets):
            ws.row_dimensions[current_row].height = 18
            use_alt = alt_idx % 2 == 1

            values = [s.gruppe, s.nummer, s.name, s.jahr, s.preis, s.growth]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center" if col in (1, 2, 4, 5, 6) else "left",
                    vertical="center",
                )
                if use_alt:
                    cell.fill = PatternFill("solid", fgColor=ROW_ALT)

            # Set-Name als Hyperlink (Spalte 3)
            name_cell = ws.cell(row=current_row, column=3)
            if s.url:
                name_cell.hyperlink = s.url
                name_cell.font = Font(
                    color="1155CC",
                    underline="single",
                    name="Calibri",
                    size=10,
                )
            else:
                name_cell.font = Font(name="Calibri", size=10)

            # Growth einfärben (Spalte 6)
            growth_cell = ws.cell(row=current_row, column=6)
            growth_val = s.growth.replace(" ", "")
            if growth_val.startswith("+"):
                growth_cell.fill = PatternFill("solid", fgColor=GREEN_BG)
                growth_cell.font = Font(bold=True, color=GREEN_FONT, size=10)
            elif growth_val.startswith("-"):
                growth_cell.fill = PatternFill("solid", fgColor=RED_BG)
                growth_cell.font = Font(bold=True, color=RED_FONT, size=10)
            elif growth_val:
                growth_cell.fill = PatternFill("solid", fgColor=NEUTRAL_BG)
                growth_cell.font = Font(bold=True, color=NEUTRAL_FONT, size=10)

            current_row += 1

    # ── Filter & Freeze ───────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:F{current_row - 1}"
    ws.freeze_panes    = "A2"

    wb.save(filepath)
    log.info("✅ Excel gespeichert: %s  (%d Sets)", filepath, sum(len(g[1]) for g in groups))


# ─── Hauptprogramm ────────────────────────────────────────────────────────────
def main():
    timestamp   = datetime.now().strftime("%Y-%m-%d_%H-%M")
    output_path = f"{OUTPUT_DIR}/brickeconomy_sets_{timestamp}.xlsx"

    session = make_scraper()

    # ── Alle Themen laden ─────────────────────────────────────────────────────
    themes = scrape_themes(session)
    if not themes:
        log.error("Keine Themen gefunden – ggf. hat sich die Seitenstruktur geändert.")
        return

    # ── Zum Testen nur ein Theme: Folgende Zeile einkommentieren ──────────────
    # themes = [t for t in themes if t["name"].lower() == "city"]

    # ── Alle Themen scrapen ───────────────────────────────────────────────────
    all_sets: list[LegoSet] = []
    for i, theme in enumerate(themes, start=1):
        log.info("─" * 60)
        log.info("Theme %d/%d: %s", i, len(themes), theme["name"])
        sets = scrape_theme(theme, session)
        all_sets.extend(sets)
        log.info("  → Theme abgeschlossen: %d Sets", len(sets))
        if i < len(themes):
            human_pause(*DELAY_THEME)

    log.info("═" * 60)
    log.info("Gesamt: %d Sets aus %d Themen", len(all_sets), len(themes))

    if all_sets:
        write_excel(all_sets, output_path)
    else:
        log.warning("Keine Sets gefunden – Excel wird nicht erstellt.")


if __name__ == "__main__":
    main()
